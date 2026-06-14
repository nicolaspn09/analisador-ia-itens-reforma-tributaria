import sys
import time
import requests
import unicodedata
import unidecode
from openpyxl import Workbook, load_workbook
from sentence_transformers import SentenceTransformer, util

sys.path.append(r"C:\rpa\Python")
from Classes.Oracle.Oracle.ConectaOracle import ConectaOracle
from Classes.Groq.Groq.acessaGroq import ExecutaGroq


def normalizar(texto):
    return unidecode.unidecode(texto.upper()).strip()


def remover_diacriticos(texto):
    nfkd = unicodedata.normalize('NFKD', texto)
    texto_sem_diacriticos = ''.join(c for c in nfkd if not unicodedata.combining(c))
    return texto_sem_diacriticos


def extrair_nome_principio(principio_completo):
    partes = principio_completo.split(" - ")
    if len(partes) >= 3:
        return partes[1].strip()
    else:
        return principio_completo.strip()


# def lista_itens_excel():
#     # Modelo pré-treinado que transforma frases em vetores
#     modelo = SentenceTransformer('paraphrase-multilingual-MiniLM-L12-v2')

#     medicamentos_reforma_analisar = medicamentos_reforma()
#     lista_principios = [normalizar(m["principio ativo"]) for m in medicamentos_reforma_analisar]
#     lista_vetores = modelo.encode(lista_principios, convert_to_tensor=True)

#     caminho_arquivo = r"C:\rpa\Python\Analisa Itens Reforma Tributaria\Analise Itens COMPANY_NAME.xlsx"
#     #Carrega a planilha principal
#     workbook = load_workbook(caminho_arquivo)
#     aba_ativa = workbook.active

#     linha = 2

#     #Itera sobre as linhas da planilha
#     for cd_mercadoria, nome_mercadoria, descricao_mercadoria, principios_ativos, unidade_de_medida, ncm, ean_venda, analise_ia, principio_anexo, item_anexo in aba_ativa.iter_rows(min_row=2, values_only=True):
#         if analise_ia == None or analise_ia == "":
#             # retorno_ia = consulta_ia(nome_mercadoria=nome_mercadoria, descricao_mercadoria=descricao_mercadoria, principios_ativos=principios_ativos, ncm=ncm)
#             # retorno_ia = consulta_ia_local(nome_mercadoria=nome_mercadoria, descricao_mercadoria=descricao_mercadoria, principios_ativos=principios_ativos, ncm=ncm)

#             principio_ativo_match = ""
#             todos_presentes = True
#             alguns_presentes = False
#             principios_do_item = []

#             if "|" in principios_ativos:
#                 principios_ativos_divididos = principios_ativos.split("|")
                
#                 for principio in principios_ativos_divididos:
#                     principio = extrair_nome_principio(principio)

#                     principio_norm = normalizar(principio)
                    
#                     # calcular embedding do princípio ativo da linha
#                     principio_vetor = modelo.encode(principio_norm, convert_to_tensor=True)
#                     # calcular similaridades
#                     similaridades = util.cos_sim(principio_vetor, lista_vetores)[0]
#                     indice_max = similaridades.argmax().item()
#                     score_max = similaridades[indice_max].item()
#                     # Define um limite de similaridade, por exemplo 0.7
#                     if score_max > 0.98:
#                         retorno_ia = "SIM"
#                         principio_ativo_match = lista_principios[indice_max]
#                         break

#             else:
#                 principio = extrair_nome_principio(principios_ativos)

#                 principio_norm = normalizar(principio)
                
#                 # calcular embedding do princípio ativo da linha
#                 principio_vetor = modelo.encode(principio_norm, convert_to_tensor=True)
#                 # calcular similaridades
#                 similaridades = util.cos_sim(principio_vetor, lista_vetores)[0]
#                 indice_max = similaridades.argmax().item()
#                 score_max = similaridades[indice_max].item()
#                 # Define um limite de similaridade, por exemplo 0.7
#                 if score_max > 0.98:
#                     retorno_ia = "SIM"
#                     principio_ativo_match = lista_principios[indice_max]
            
#             if "SIM" in retorno_ia:
#                 preenche_informacao_excel(aba_ativa, linha, retorno_ia, principio_ativo_match, (indice_max + 2))
#             else:
#                 preenche_informacao_excel(aba_ativa, linha, retorno_ia, principio_ativo_match, "")
#             workbook.save(caminho_arquivo)
#             linha += 1
#             time.sleep(2)
#         else:
#             linha += 1

#     workbook.close()


def lista_itens_excel():
    # Modelo pré-treinado que transforma frases em vetores
    modelo = SentenceTransformer('paraphrase-multilingual-MiniLM-L12-v2')

    medicamentos_reforma_analisar = medicamentos_reforma()
    lista_principios_originais = [m["principio ativo"] for m in medicamentos_reforma_analisar]
    lista_principios_normalizados = [normalizar(m["principio ativo"]) for m in medicamentos_reforma_analisar]
    lista_vetores = modelo.encode(lista_principios_normalizados, convert_to_tensor=True)

    caminho_arquivo = r"C:\rpa\Python\Analisa Itens Reforma Tributaria\Analise Itens COMPANY_NAME.xlsx"
    # Carrega a planilha principal
    workbook = load_workbook(caminho_arquivo)
    aba_ativa = workbook.active

    linha = 2

    # Itera sobre as linhas da planilha
    for cd_mercadoria, nome_mercadoria, descricao_mercadoria, principios_ativos, unidade_de_medida, ncm, ean_venda, analise_ia_excel, principio_anexo_excel, item_anexo_excel in aba_ativa.iter_rows(min_row=2, values_only=True):
        if analise_ia_excel is None or analise_ia_excel == "":
            retorno_ia = "NÃO"
            principios_ativos_encontrados = []
            indices_anexo_encontrados = []
            todos_presentes = True
            alguns_presentes = False

            principios_do_item = []
            if principios_ativos:
                principios_do_item = [extrair_nome_principio(p).strip() for p in principios_ativos.split("|")]
            
            # Divide princípios ativos com "+" e verifica todos os componentes
            principios_para_analisar = []
            for p_item in principios_do_item:
                if "+" in p_item:
                    principios_para_analisar.extend([s.strip() for s in p_item.split("+")])
                else:
                    principios_para_analisar.append(p_item)

            if principios_para_analisar:
                for principio_analise in principios_para_analisar:
                    principio_norm = normalizar(principio_analise)
                    
                    # Calcular embedding do princípio ativo da linha
                    principio_vetor = modelo.encode(principio_norm, convert_to_tensor=True)
                    # Calcular similaridades
                    similaridades = util.cos_sim(principio_vetor, lista_vetores)[0]
                    indice_max = similaridades.argmax().item()
                    score_max = similaridades[indice_max].item()

                    if score_max > 0.98:
                        alguns_presentes = True
                        # Adiciona o princípio ativo original da lista de referência e seu índice
                        principios_ativos_encontrados.append(lista_principios_originais[indice_max])
                        indices_anexo_encontrados.append(str(indice_max + 2)) # +2 para corresponder à linha do Excel
                        
                        # Verifica se o princípio ativo original da lista (que pode ter "+") contém todos os princípios do item
                        if not all(normalizar(comp) in lista_principios_normalizados[indice_max] for comp in [normalizar(s.strip()) for s in principio_analise.split('+')]):
                            todos_presentes = False
                    else:
                        todos_presentes = False
                
                if todos_presentes and alguns_presentes:
                    retorno_ia = "SIM"
                elif alguns_presentes and todos_presentes and "|" in principios_ativos:
                    retorno_ia = "REVISAR"
                else:
                    retorno_ia = "NÃO" # Caso nenhum princípio ativo do item seja encontrado

            # Formata as strings para preencher o Excel
            principios_match_str = ", ".join(sorted(list(set(principios_ativos_encontrados)))) # Remove duplicatas e ordena
            indices_anexo_str = ", ".join(sorted(list(set(indices_anexo_encontrados)), key=int)) # Remove duplicatas e ordena numericamente

            if "SIM" in retorno_ia or "REVISAR" in retorno_ia:
                preenche_informacao_excel(aba_ativa, linha, retorno_ia, principios_match_str, indices_anexo_str)
            else:
                preenche_informacao_excel(aba_ativa, linha, retorno_ia, "", "")
            
            workbook.save(caminho_arquivo)
            linha += 1
            time.sleep(2)
        else:
            linha += 1

    workbook.close()


def preenche_informacao_excel(aba_ativa, linha, informacao_ia, principio_ativo, contador_principio):
    aba_ativa.cell(row=linha, column=8).value = informacao_ia
    aba_ativa.cell(row=linha, column=9).value = principio_ativo
    aba_ativa.cell(row=linha, column=10).value = contador_principio


def consulta_ia(nome_mercadoria, descricao_mercadoria, principios_ativos, ncm):
    medicamentos = medicamentos_reforma()

    system = f"""
    Você é um assistente fiscal, e está trabalhando na análise dos itens da reforma tributária.

    Vou lhe enviar a princípios ativos que estão na isenção de 0%.
    Seu trabalho é analisar a mercadoria que vou te passar e me retornar se ele terá isenção ou não.
    Retorno esperado: SIM ou NÃO e a explicação do porque ter ou não ter isenção.
    Obs.: Pode acontecer uma disparidade nos princípios ativos, mas os nomes não serão muito diferentes da lista e da mercadoria pra ser analisada.

    Lista de princípios ativos presentes na isenção: {medicamentos}
    """

    user = f"""Mercadoria para ser analisada: 
    Nome da mercadoria: {nome_mercadoria},
    Descrição da mercadoria: {descricao_mercadoria},
    Princípios ativos da mercadoria: {principios_ativos},
    NCM da mercadoria: {ncm}
    """

    chave = "'SECRET_REMOVED_BY_AI'"
    modelo = "meta-llama/llama-4-scout-17b-16e-instruct"
    # modelo = "meta-llama/llama-guard-4-12b"
    # modelo = "gemma2-9b-it"


    groq = ExecutaGroq(chave=chave, model=modelo, prompt=system, texto=user)
    try:
        resultado = groq.executa_groq()
    except:
        time.sleep(60)
        resultado = groq.executa_groq()

    if "</think>" in resultado:
        resultado = resultado.split("/<think>")[1]

    return str(resultado).replace("\n", " ").strip()


def consulta_ia_local(nome_mercadoria, descricao_mercadoria, principios_ativos, ncm):
    medicamentos = medicamentos_reforma()

    system = f"""
    Você é um assistente fiscal, e está trabalhando na análise dos itens da reforma tributária.

    Vou lhe enviar princípios ativos que estão na isenção de 0%.
    Seu trabalho é analisar a mercadoria que vou te passar e me retornar se ela terá isenção ou não, com base principalmente no princípio ou princípios ativos dessa mercadoria que te passei.

    Retorno esperado: SIM ou NÃO e a explicação do porque ter ou não ter isenção.
    Obs.: Pode acontecer uma disparidade nos princípios ativos da mercadoria a ser analisada, mas os nomes não serão muito diferentes da lista e da mercadoria para ser analisada.

    Lista de princípios ativos presentes na isenção: {medicamentos}
    """

    user = f"""Mercadoria para ser analisada: 
    Nome da mercadoria: {nome_mercadoria},
    Descrição da mercadoria: {descricao_mercadoria},
    Princípios ativos da mercadoria: {principios_ativos},
    NCM da mercadoria: {ncm}
    """

    # system = "Você é um assistente virtual"
    # user = "Me fale sobre Alan Turing"

    prompt = f"Prompt de orientação: {system}.\n\nAnalise essa mercadoria e veja se ela está presente na lista de isenções enviados anteriormente: {user}. Use o o comando 'Retorno esperado' do prompt de orientação que eu te enviei para me responder"
    payload = {
        "model": "deepseek-r1:7b", # gemma3:4b # llama3.1:8b
        "prompt": prompt,
        "stream": False
    }

    try:
        r = requests.post("http://localhost:11434/api/generate", json=payload, timeout=240)
        r.raise_for_status()
        resultado = r.json()["response"]
    except:
        time.sleep(60)
        r = requests.post("http://localhost:11434/api/generate", json=payload, timeout=240)
        r.raise_for_status()
        resultado = r.json()["response"]

    if "</think>" in str(resultado):
        resultado = str(resultado).split("</think>")[1]

    if "Retorno esperado:" in str(resultado):
        resultado = str(resultado).replace("Retorno esperado:", "").strip()

    return resultado.replace("\n", " ").strip()


def medicamentos_reforma():
    medicamentos = [
        {"principio ativo": "ABACAVIR"},
        {"principio ativo": "ABEMACICLIBE"},
        {"principio ativo": "ACALABRUTINIBE"},
        {"principio ativo": "ACEPONATO DE METILPREDNISOLONA"},
        {"principio ativo": "ACETATO DE ABIRATERONA"},
        {"principio ativo": "ACETATO DE CIPROTERONA"},
        {"principio ativo": "ACETATO DE DEGARELIX"},
        {"principio ativo": "ACETATO DE GOSSERRELINA"},
        {"principio ativo": "ACETATO DE LEUPRORRELINA"},
        {"principio ativo": "ACETATO DE MEGESTROL"},
        {"principio ativo": "ACETATO DE OCTREOTIDA"},
        {"principio ativo": "ACETATO DE TRIPTORRELINA"},
        {"principio ativo": "ACETATO DESMOPRESSINA"},
        {"principio ativo": "ÁCIDO FOLÍNICO (FÓLICO)"},
        {"principio ativo": "ÁCIDO TRANEXÂMICO"},
        {"principio ativo": "ÁCIDO ZOLEDRÔNICO"},
        {"principio ativo": "ACITRETINA"},
        {"principio ativo": "AFLIBERCEPTE"},
        {"principio ativo": "ALBINTERFERONA ALFA-2B"},
        {"principio ativo": "ALBUMINA HUMANA"},
        {"principio ativo": "ALENDRONATO DE SÓDIO"},
        {"principio ativo": "ALENTUZUMABE"},
        {"principio ativo": "ALFA-ALGLICOSIDASE"},
        {"principio ativo": "ALFAELOSULFASE"},
        {"principio ativo": "ALFAEPOETINA"},
        {"principio ativo": "ALFAINTERFERONA"},
        {"principio ativo": "ALFAPEGINTERFERONA 2A"},
        {"principio ativo": "ALFAPEGINTERFERONA 2B"},
        {"principio ativo": "ALFATIROTROPINA"},
        {"principio ativo": "ALFAVESTRONIDASE"},
        {"principio ativo": "ALPELISIBE"},
        {"principio ativo": "ALTEPLASE"},
        {"principio ativo": "AMBRISENTANA"},
        {"principio ativo": "AMIFOSTINA"},
        {"principio ativo": "ANASTROZOL"},
        {"principio ativo": "ANFOTERICINA B"},
        {"principio ativo": "ANFOTERICINA B EM LIPOSSOMAS"},
        {"principio ativo": "ANTIMONIAL PENTAVALENTE"},
        {"principio ativo": "APALUTAMIDA"},
        {"principio ativo": "APREPITANTO"},
        {"principio ativo": "ARTEMÉTER"},
        {"principio ativo": "ARTEMÉTER + LUMEFANTRINA"},
        {"principio ativo": "ARTESUNATO + CLORIDRATO MEFLOQUINA"},
        {"principio ativo": "ARTESUNATO DE SÓDIO"},
        {"principio ativo": "ASPARAGINASE"},
        {"principio ativo": "ATENOLOL"},
        {"principio ativo": "ATEZOLIZUMABE"},
        {"principio ativo": "AVELUMABE"},
        {"principio ativo": "AXITINIBE"},
        {"principio ativo": "AZACITIDINA"},
        {"principio ativo": "AZATIOPRINA"},
        {"principio ativo": "BARICITINIBE"},
        {"principio ativo": "BENZONIDAZOL"},
        {"principio ativo": "BESILATO DE ANLODIPINO"},
        {"principio ativo": "BETAEPOETINA"},
        {"principio ativo": "BEVACIZUMABE"},
        {"principio ativo": "BICALUTAMIDA"},
        {"principio ativo": "BIOTINA"},
        {"principio ativo": "BLINATUMOMABE"},
        {"principio ativo": "BORTEZOMIBE"},
        {"principio ativo": "BRENTUXIMABE VEDOTINA"},
        {"principio ativo": "BRIGATINIBE"},
        {"principio ativo": "BROMETO DE IPRATRÓPIO"},
        {"principio ativo": "BUDESONIDA"},
        {"principio ativo": "BUROSUMABE"},
        {"principio ativo": "BUSSULFANO"},
        {"principio ativo": "CABAZITAXEL"},
        {"principio ativo": "CAPECITABINA"},
        {"principio ativo": "CARBIDOPA + LEVODOPA"},
        {"principio ativo": "CARBOPLATINA"},
        {"principio ativo": "CARFILZOMIBE"},
        {"principio ativo": "CARMUSTINA"},
        {"principio ativo": "CEFALOTINA"},
        {"principio ativo": "CEFOXITINA"},
        {"principio ativo": "CEFTAZIDIMA"},
        {"principio ativo": "CELECOXIBE"},
        {"principio ativo": "CETUXIMABE"},
        {"principio ativo": "CICLOFOSFAMIDA"},
        {"principio ativo": "CILASTATINA SÓDICA + IMIPENEM"},
        {"principio ativo": "CISPLATINA"},
        {"principio ativo": "CITARABINA"},
        {"principio ativo": "CITRATO DE IXAZOMIBE"},
        {"principio ativo": "CITRATO DE TAMOXIFENO"},
        {"principio ativo": "CLADRIBINA"},
        {"principio ativo": "CLODRONATO DISSÓDICO"},
        {"principio ativo": "CLOFAZIMINA"},
        {"principio ativo": "CLORAMBUCILA"},
        {"principio ativo": "CLORETO DE RÁDIO (223 RA)"},
        {"principio ativo": "CLORETO DE SÓDIO"},
        {"principio ativo": "CLORETO DE SUXAMETÔNIO"},
        {"principio ativo": "CLORIDRATO DE ALECTINIBE"},
        {"principio ativo": "CLORIDRATO DE ALFENTANILA MONOIDRATADA"},
        {"principio ativo": "CLORIDRATO DE AMINOLEVULINATO DE METILA"},
        {"principio ativo": "CLORIDRATO DE CINACALCETE"},
        {"principio ativo": "CLORIDRATO DE DAUNORRUBICINA"},
        {"principio ativo": "CLORIDRATO DE DOBUTAMINA"},
        {"principio ativo": "CLORIDRATO DE DOXORRUBICINA"},
        {"principio ativo": "CLORIDRATO DE EPIRRUBICINA"},
        {"principio ativo": "CLORIDRATO DE ERLOTINIBE"},
        {"principio ativo": "CLORIDRATO DE FINGOLIMODE"},
        {"principio ativo": "CLORIDRATO DE GENCITABINA"},
        {"principio ativo": "CLORIDRATO DE GRANISSETRONA"},
        {"principio ativo": "CLORIDRATO DE IDARRUBICINA"},
        {"principio ativo": "CLORIDRATO DE IRINOTECANO"},
        {"principio ativo": "CLORIDRATO DE IRINOTECANO TRI-HIDRATADO"},
        {"principio ativo": "CLORIDRATO DE METOCLOPRAMIDA"},
        {"principio ativo": "CLORIDRATO DE MITOXANTRONA"},
        {"principio ativo": "CLORIDRATO DE PALONOSETRONA"},
        {"principio ativo": "CLORIDRATO DE PAZOPANIBE"},
        {"principio ativo": "CLORIDRATO DE PIRIDOXINA"},
        {"principio ativo": "CLORIDRATO DE PONATINIBE"},
        {"principio ativo": "CLORIDRATO DE TOPOTECANA"},
        {"principio ativo": "CLORIDRATO DE ZIPRASIDONA MONOIDRATADO"},
        {"principio ativo": "COMPLEXO PROTROMBÍNICO PARCIALMENTE ATIVADO"},
        {"principio ativo": "CRIZOTINIBE"},
        {"principio ativo": "DACARBAZINA"},
        {"principio ativo": "DAPAGLIFLOZINA"},
        {"principio ativo": "DARATUMUMABE"},
        {"principio ativo": "DAROLUTAMIDA"},
        {"principio ativo": "DASATINIBE"},
        {"principio ativo": "DECITABINA"},
        {"principio ativo": "DEFERASIROX"},
        {"principio ativo": "DENOSUMABE"},
        {"principio ativo": "DEXAMETASONA"},
        {"principio ativo": "DIASPARTATO DE PASIREOTIDA"},
        {"principio ativo": "DIAZEPAM"},
        {"principio ativo": "DICLORIDRATO DE DACLATASVIR"},
        {"principio ativo": "DICLORIDRATO DE PRAMIPEXOL MONOIDRATADO"},
        {"principio ativo": "DICLORIDRATO DE QUININA"},
        {"principio ativo": "DICLORIDRATO DE SAPROPTERINA"},
        {"principio ativo": "DIDANOSINA"},
        {"principio ativo": "DIETILESTILBESTROL"},
        {"principio ativo": "DIFOSFATO DE CLOROQUINA"},
        {"principio ativo": "DIMALEATO DE AFATINIBE"},
        {"principio ativo": "DIMETILSULFÓXIDO DE TRAMETINIBE"},
        {"principio ativo": "DITARTARATO DE VINORELBINA"},
        {"principio ativo": "DOCETAXEL"},
        {"principio ativo": "DOCETAXEL TRI-HIDRATADO"},
        {"principio ativo": "DOLUTEGRAVIR SÓDICO"},
        {"principio ativo": "DOXICICLINA MONOIDRATADA"},
        {"principio ativo": "DURVALUMABE"},
        {"principio ativo": "ECULIZUMABE"},
        {"principio ativo": "EFAVIRENZ"},
        {"principio ativo": "ELEXACAFTOR"},
        {"principio ativo": "ELOTUZUMABE"},
        {"principio ativo": "ELTROMBOPAGUE OLAMINA"},
        {"principio ativo": "EMBONATO DE TRIPTORRELINA"},
        {"principio ativo": "EMICIZUMABE"},
        {"principio ativo": "EMTRICITABINA"},
        {"principio ativo": "ENANTATO DE NORETISTERONA + VALERATO DE ESTRADIOL"},
        {"principio ativo": "ENFLURANO"},
        {"principio ativo": "ENFUVIRTIDA"},
        {"principio ativo": "ENTRICITABINA"},
        {"principio ativo": "ENTRICITABINA + FUMARATO TENOFOVIR DESOPROXILA"},
        {"principio ativo": "ENZALUTAMIDA"},
        {"principio ativo": "ERDAFITINIBE"},
        {"principio ativo": "ESILATO DE NINTEDANIBE"},
        {"principio ativo": "ESPIRONOLACTONA"},
        {"principio ativo": "ESTAVUDINA"},
        {"principio ativo": "ETINILESTRADIOL + LEVONORGESTREL"},
        {"principio ativo": "ETOMIDATO"},
        {"principio ativo": "ETOPOSIDEO"},
        {"principio ativo": "ETRAVIRINA"},
        {"principio ativo": "EVEROLIMO"},
        {"principio ativo": "EXEMESTANO"},
        {"principio ativo": "FATOR IX DE COAGULAÇÃO"},
        {"principio ativo": "FATOR VII DE COAGULAÇÃO ATIVADO RECOMBINANTE"},
        {"principio ativo": "FATOR VIII DE COAGULAÇÃO"},
        {"principio ativo": "FATOR VIII DE COAGULAÇÃO CONTENDO FATOR DE VON WILLEBRAND"},
        {"principio ativo": "FATOR VIII DE COAGULAÇÃO RECOMBINANTE"},
        {"principio ativo": "FENTANILA"},
        {"principio ativo": "FILGRASTIM"},
        {"principio ativo": "FLUORURACILA"},
        {"principio ativo": "FOLINATO DE CÁLCIO"},
        {"principio ativo": "FOSAMPRENAVIR CÁLCICO"},
        {"principio ativo": "FOSFATO DE FLUDARABINA"},
        {"principio ativo": "FOSFATO DE OSELTAMIVIR"},
        {"principio ativo": "FOSFATO DE RUXOLITINIBE"},
        {"principio ativo": "FOSFATO DE SITAGLIPTINA"},
        {"principio ativo": "FOTEMUSTINA"},
        {"principio ativo": "FULVESTRANTO"},
        {"principio ativo": "FUMARATO DE DIMETILA"},
        {"principio ativo": "FUMARATO DE TENOFOVIR DESOPROXILA"},
        {"principio ativo": "FUROSEMIDA"},
        {"principio ativo": "GALSULFASE"},
        {"principio ativo": "GANCICLOVIR SÓDICO"},
        {"principio ativo": "GEFITINIBE"},
        {"principio ativo": "GLICOSE"},
        {"principio ativo": "GOLIMUMABE"},
        {"principio ativo": "GOSSERRELINA"},
        {"principio ativo": "GRANISETRON"},
        {"principio ativo": "HALOPERIDOL"},
        {"principio ativo": "HIDROXIUREIA"},
        {"principio ativo": "HIPOCLORITO DE SÓDIO"},
        {"principio ativo": "IBANDRONATO SÓDIO"},
        {"principio ativo": "IBRUTINIBE"},
        {"principio ativo": "IDARRUBICINA"},
        {"principio ativo": "IDURSULFASE"},
        {"principio ativo": "IFOSFAMIDA"},
        {"principio ativo": "IMUNOGLOBULINA ANTI-HEPATITE B"},
        {"principio ativo": "IMUNOGLOBULINA ANTIRRÁBICA"},
        {"principio ativo": "IMUNOGLOBULINA ANTITETÂNICA"},
        {"principio ativo": "INSULINA GLARGINA"},
        {"principio ativo": "INSULINA HUMANA"},
        {"principio ativo": "INTERFERON ALFA-2A E INTERFERON ALFA-2B"},
        {"principio ativo": "IOPAMIDOL"},
        {"principio ativo": "IPILIMUMABE"},
        {"principio ativo": "ISETIONATO DE PENTAMIDINA"},
        {"principio ativo": "ISOFLURANO"},
        {"principio ativo": "ISOTRETINOÍNA"},
        {"principio ativo": "IVACAFTOR"},
        {"principio ativo": "LAMIVUDINA + ZIDOVUDINA"},
        {"principio ativo": "LETROZOL"},
        {"principio ativo": "LEVETIRACETAM"},
        {"principio ativo": "LIDOCAÍNA"},
        {"principio ativo": "LINEZOLIDA"},
        {"principio ativo": "LIPEGFILGRASTIM"},
        {"principio ativo": "LOPINAVIR + RITONAVIR"},
        {"principio ativo": "LOSARTANA POTÁSSICA"},
        {"principio ativo": "LUMACAFTOR"},
        {"principio ativo": "MALEATO DE ACALABRUTINIBE MONOIDRATADO"},
        {"principio ativo": "MALEATO DE SUNITINIBE"},
        {"principio ativo": "MALEATO DE TIMOLOL"},
        {"principio ativo": "MARAVIROQUE"},
        {"principio ativo": "MEPOLIZUMABE"},
        {"principio ativo": "MERCAPTOPURINA"},
        {"principio ativo": "MESILATO DE DABRAFENIBE"},
        {"principio ativo": "MESILATO DE DESFERROXAMINA"},
        {"principio ativo": "MESILATO DE IMATINIBE"},
        {"principio ativo": "MESILATO DE NELFINAVIR"},
        {"principio ativo": "MESILATO DE OSIMERTINIBE"},
        {"principio ativo": "MESILATO DE RASAGILINA"},
        {"principio ativo": "MESNA"},
        {"principio ativo": "METILPREDNISOLONA"},
        {"principio ativo": "METOTREXATO"},
        {"principio ativo": "METOTREXATO DE SÓDIO"},
        {"principio ativo": "MICOFENOLATO DE MOFETILA"},
        {"principio ativo": "MICOFENOLATO DE SÓDIO"},
        {"principio ativo": "MIDAZOLAM"},
        {"principio ativo": "MIDOSTAURINA"},
        {"principio ativo": "MIFAMURTIDA"},
        {"principio ativo": "MITOMICINA"},
        {"principio ativo": "MITOTANO"},
        {"principio ativo": "NEVIRAPINA"},
        {"principio ativo": "NILOTINIBE"},
        {"principio ativo": "NITRENDIPINO"},
        {"principio ativo": "NIVOLUMABE"},
        {"principio ativo": "NUSINERSENA"},
        {"principio ativo": "OCRELIZUMABE"},
        {"principio ativo": "OCTREOTIDA"},
        {"principio ativo": "OLAPARIBE"},
        {"principio ativo": "OLARATUMABE"},
        {"principio ativo": "ONASEMNOGENO ABEPARVOVEQUE"},
        {"principio ativo": "OXALIPLATINA"},
        {"principio ativo": "PACLITAXEL"},
        {"principio ativo": "PALBOCICLIBE"},
        {"principio ativo": "PAMIDRONATO DISSÓDICO"},
        {"principio ativo": "PAMOATO DE PASIREOTIDA"},
        {"principio ativo": "PANCREATINA"},
        {"principio ativo": "PANITUMUMABE"},
        {"principio ativo": "PEG INTERFERON ALFA-2B"},
        {"principio ativo": "PEG INTERFERON ALFA-2A"},
        {"principio ativo": "PEGASPARGASE"},
        {"principio ativo": "PEGFILGRASTIM"},
        {"principio ativo": "PEMETREXEDE DISSÓDICO"},
        {"principio ativo": "PEMETREXEDE DISSÓDICO HEMIPENTAIDRATADO"},
        {"principio ativo": "PEMETREXEDE DISSÓDICO HEPTAIDRATADO"},
        {"principio ativo": "PERTUZUMABE"},
        {"principio ativo": "PIOGLITAZONA"},
        {"principio ativo": "PIRAZINAMIDA + RIFAMPICINA + CLORIDRATO DE ETAMBUTOL + ISONIAZIDA"},
        {"principio ativo": "PLERIXAFOR"},
        {"principio ativo": "PRAZIQUANTEL"},
        {"principio ativo": "PREDNISOLONA"},
        {"principio ativo": "PREGABALINA"},
        {"principio ativo": "PROPOFOL"},
        {"principio ativo": "QUININA"},
        {"principio ativo": "RABEPRAZOL SÓDICO"},
        {"principio ativo": "RALTEGRAVIR"},
        {"principio ativo": "RAMUCIRUMABE"},
        {"principio ativo": "RASBURICASE"},
        {"principio ativo": "REGORAFENIBE"},
        {"principio ativo": "RIBAVIRINA"},
        {"principio ativo": "RIFAMPICINA + ISONIAZIDA"},
        {"principio ativo": "RILUZOL"},
        {"principio ativo": "RISANQUIZUMABE"},
        {"principio ativo": "RISDIPLAM"},
        {"principio ativo": "RISPERIDONA"},
        {"principio ativo": "RITONAVIR"},
        {"principio ativo": "RITUXIMABE"},
        {"principio ativo": "SACUBITRIL VALSARTANA SÓDICA HIDRATADA"},
        {"principio ativo": "SAQUINAVIR"},
        {"principio ativo": "SAXAGLIPTINA"},
        {"principio ativo": "SECUQUINUMABE"},
        {"principio ativo": "SELEXIPAGUE"},
        {"principio ativo": "SINVASTATINA"},
        {"principio ativo": "SOFOSBUVIR"},
        {"principio ativo": "SOMATROPINA"},
        {"principio ativo": "SORO ANTIARACNÍDICO (LOXOSCELES, PHONEUTRIA E TITYUS)"},
        {"principio ativo": "SORO ANTIBOTRÓPICO (PENTAVALENTE)"},
        {"principio ativo": "SORO ANTIBOTRÓPICO (PENTAVALENTE) E ANTICROTÁLICO"},
        {"principio ativo": "SORO ANTIBOTRÓPICO (PENTAVALENTE) E ANTILAQUÉTICO"},
        {"principio ativo": "SORO ANTIBOTULÍNICO AB (BIVALENTE)"},
        {"principio ativo": "SORO ANTICROTÁLICO"},
        {"principio ativo": "SORO ANTIDIFTÉRICO"},
        {"principio ativo": "SORO ANTIELAPÍDICO (BIVALENTE)"},
        {"principio ativo": "SORO ANTIESCORPIÔNICO"},
        {"principio ativo": "SORO ANTILONÔMICO"},
        {"principio ativo": "SORO ANTILOXOSCÉLICO (TRIVALENTE)"},
        {"principio ativo": "SORO ANTIRRÁBICO"},
        {"principio ativo": "SORO ANTITETÂNICO"},
        {"principio ativo": "SUCCINATO DE METOPROLOL"},
        {"principio ativo": "SUCCINATO DE RIBOCICLIBE"},
        {"principio ativo": "SUCCINATO SÓDICO DE HIDROCORTISONA"},
        {"principio ativo": "SULFADIAZINA"},
        {"principio ativo": "SULFAMETOXAZOL + TRIMETROPINA"},
        {"principio ativo": "SULFATO DE ABACAVIR"},
        {"principio ativo": "SULFATO DE ATAZANAVIR"},
        {"principio ativo": "SULFATO DE BLEOMICINA"},
        {"principio ativo": "SULFATO DE INDINAVIR"},
        {"principio ativo": "SULFATO DE LAROTRECTINIBE"},
        {"principio ativo": "SULFATO DE MORFINA"},
        {"principio ativo": "SULFATO DE MORFINA PENTAIDRATADO"},
        {"principio ativo": "SULFATO DE QUININA"},
        {"principio ativo": "SULFATO DE VINCRISTINA"},
        {"principio ativo": "TACROLIMO"},
        {"principio ativo": "TAFAMIDIS MEGLUMINA"},
        {"principio ativo": "TAMOXIFENO"},
        {"principio ativo": "TARTARATO DE VARENICLINA"},
        {"principio ativo": "TARTARATO DE VINORELBINA"},
        {"principio ativo": "TEMOZOLOMIDA"},
        {"principio ativo": "TENECTEPLASE"},
        {"principio ativo": "TENIPOSIDEO"},
        {"principio ativo": "TENOFOVIR"},
        {"principio ativo": "TENSIROLIMO"},
        {"principio ativo": "TERIFLUNOMIDA"},
        {"principio ativo": "TERIZIDONA"},
        {"principio ativo": "TETRACICLINA"},
        {"principio ativo": "TEZACAFTOR"},
        {"principio ativo": "TIOGUANINA"},
        {"principio ativo": "TIPRANAVIR"},
        {"principio ativo": "TOCILIZUMABE"},
        {"principio ativo": "TOSILATO DE SORAFENIBE"},
        {"principio ativo": "TRASTUZUMABE"},
        {"principio ativo": "TRIÓXIDO DE ARSÊNIO"},
        {"principio ativo": "TRIPTORRELINA"},
        {"principio ativo": "UPADACITINIBE HEMI-HIDRATADO"},
        {"principio ativo": "VANCOMICINA"},
        {"principio ativo": "VANDETANIBE"},
        {"principio ativo": "VEDOLIZUMABE"},
        {"principio ativo": "VIMBLASTINA"},
        {"principio ativo": "VINCRISTINA"},
        {"principio ativo": "VINFLUNINA"},
        {"principio ativo": "VINORELBINA"},
        {"principio ativo": "ZIAGENAVIR"},
        {"principio ativo": "ZIDOVUDINA"},
        {"principio ativo": "VACINA ADSORVIDA DIFTERIA E TÉTANO"},
        {"principio ativo": "VACINA ADSORVIDA DIFTERIA, TÉTANO E PERTUSSIS"},
        {"principio ativo": "VACINA ADSORVIDA DIFTERIA, TÉTANO E PERTUSSIS (ACELULAR)"},
        {"principio ativo": "VACINA ADSORVIDA DIFTERIA, TÉTANO, PERTUSSIS, HEPATITE B (RECOMBINANTE) E HAEMOPHILUS INFLUENZAE B (CONJUGADA)"},
        {"principio ativo": "VACINA ADSORVIDA HEPATITE A (INATIVADA)"},
        {"principio ativo": "VACINA BCG"},
        {"principio ativo": "VACINA CÓLERA (INATIVADA)"},
        {"principio ativo": "VACINA COVID-19"},
        {"principio ativo": "VACINA DENGUE 1, 2, 3 E 4"},
        {"principio ativo": "VACINA FEBRE AMARELA (ATENUADA)"},
        {"principio ativo": "VACINA FEBRE TIFÓIDE (POLISSACARÍDICA)"},
        {"principio ativo": "VACINA HAEMOPHILUS INFLUENZAE B (CONJUGADA)"},
        {"principio ativo": "VACINA HEPATITE B (RECOMBINANTE)"},
        {"principio ativo": "VACINA INFLUENZA TRIVALENTE (FRAGMENTADA, INATIVADA)"},
        {"principio ativo": "VACINA MENINGOCÓCICA ACWY (CONJUGADA)"},
        {"principio ativo": "VACINA MENINGOCÓCICA C (CONJUGADA)"},
        {"principio ativo": "VACINA PAPILOMAVÍRUS HUMANO 6, 11, 16 E 18 (RECOMBINANTE)"},
        {"principio ativo": "VACINA PNEUMOCÓCICA 10-VALENTE (CONJUGADA)"},
        {"principio ativo": "VACINA PNEUMOCÓCICA 13-VALENTE (CONJUGADA)"},
        {"principio ativo": "VACINA PNEUMOCÓCICA 23-VALENTE (POLISSACARÍDICA)"},
        {"principio ativo": "VACINA POLIOMIELITE 1 E 3 (ATENUADA)"},
        {"principio ativo": "VACINA POLIOMIELITE 1, 2 E 3 (INATIVADA)"},
        {"principio ativo": "VACINA RAIVA (INATIVADA)"},
        {"principio ativo": "VACINA ROTAVÍRUS HUMANO G1P 8 (ATENUADA)"},
        {"principio ativo": "VACINA SARAMPO, CAXUMBA, RUBÉOLA"},
        {"principio ativo": "VACINA SARAMPO, CAXUMBA, RUBÉOLA E VARICELA (ATENUADA)"},
        {"principio ativo": "VACINA VARICELA (ATENUADA)"}
    ]

    return medicamentos

    
lista_itens_excel()