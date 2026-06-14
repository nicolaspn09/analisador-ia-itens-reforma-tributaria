import re
import time
import requests
import unicodedata
import unidecode
from openpyxl import load_workbook
from geraTabelasSheetsIsencao import GeraTabelasSheetsIsencao
from sentence_transformers import SentenceTransformer, CrossEncoder, util


# --- CONFIGURAÇÃO DOS MODELOS ---
# Bi-encoder para similaridade rápida (usado em medicamentos/princípios)
modelo_bi = None
# Cross-encoder para "interpretação" de relevância (usado em descrições complexas)
modelo_cross = None

def carregar_modelos():
    global modelo_bi, modelo_cross
    print("Carregando modelo Bi-Encoder (Medicamentos)...")
    modelo_bi = SentenceTransformer('paraphrase-multilingual-MiniLM-L12-v2')
    
    print("Carregando modelo Cross-Encoder (Interpretação de Descrições)...")
    # Este modelo é especifico para ver se Texto B responde/bate com Texto A em Português
    modelo_cross = CrossEncoder('cross-encoder/mmarco-mMiniLMv2-L12-H384-v1')


def normalizar(texto):
    if not texto: 
        return ""
    return unidecode.unidecode(str(texto).upper()).strip()


def limpar_ncm(ncm):
    """
    Remove pontos, espaços e caracteres não numéricos.
    Ex: '9018.11.00' -> '90181100'
    """
    if not ncm: 
        return ""
    return "".join(filter(str.isdigit, str(ncm)))


def remover_diacriticos(texto):
    nfkd = unicodedata.normalize('NFKD', texto)
    texto_sem_diacriticos = ''.join(c for c in nfkd if not unicodedata.combining(c))
    return texto_sem_diacriticos


def extrair_ncms_excecao(texto_descricao):
    """
    Procura por palavras de exclusão e extrai NCMs listados após elas.
    """
    texto_norm = normalizar(texto_descricao).lower()
    gatilhos = ["exceto", "excecao", "salvo", "excluindo", "excluidos"]
    
    inicio_excecao = -1
    for gatilho in gatilhos:
        idx = texto_norm.find(gatilho)
        if idx != -1:
            inicio_excecao = idx
        break
            
    if inicio_excecao == -1:
        return []
    
    # Pega o texto da exceção em diante
    texto_excecao = texto_descricao[inicio_excecao:]
    
    # Regex robusta: captura sequências numéricas que pareçam NCMs (com ou sem pontos)
    # Ex: captura 9018, 9018.11, 9018.11.00
    matches = re.findall(r'\b\d{4,8}\b|\b\d{4}\.\d{2}\b|\b\d{4}\.\d{2}\.\d{2}\b', texto_excecao)
    
    # Retorna lista limpa (apenas números)
    return [limpar_ncm(m) for m in matches]


def extrair_nome_principio(principio_completo):
    partes = principio_completo.split(" - ")
    if len(partes) >= 3:
        return partes[1].strip()
    else:
        return principio_completo.strip()
    

def preenche_informacao_excel(aba_ativa, linha, retorno_ia, principio_ativo, item_anexo, isento_convenio):
    aba_ativa.cell(row=linha, column=11).value = retorno_ia
    aba_ativa.cell(row=linha, column=12).value = principio_ativo
    aba_ativa.cell(row=linha, column=13).value = item_anexo
    aba_ativa.cell(row=linha, column=14).value = isento_convenio


# def comparar_descricao_semantica(texto_item, texto_tabela, threshold=0.97):
#     # Ignora o texto de exceção para a comparação semântica
#     # Ex: Se a tabela diz "Aparelhos X, exceto NCM Y", comparamos apenas "Aparelhos X"
#     texto_tabela_limpo = re.split(r'exceto|exceção|salvo', texto_tabela, flags=re.IGNORECASE)[0].strip()
    
#     texto_item_norm = normalizar(texto_item)
#     texto_tabela_norm = normalizar(texto_tabela_limpo)

#     if texto_tabela_norm == texto_item_norm: 
#         return True, 1.0
    
#     # Contenção: Só aceitamos se a palavra chave da tabela estiver no item
#     # Mas cuidado: Se a tabela for "LUVAS", e o item "LUVAS DE ORO", ok.
#     if texto_tabela_norm in texto_item_norm: 
#         return True, 0.99

#     if modelo:
#         try:
#             v1 = modelo.encode(texto_item_norm, convert_to_tensor=True)
#             v2 = modelo.encode(texto_tabela_norm, convert_to_tensor=True)
#             sim = util.cos_sim(v1, v2)[0][0].item()
#             if sim > threshold: 
#                 return True, sim
#         except: 
#             pass

#     return False, 0.0


# --- LÓGICA 1: MEDICAMENTOS (Bi-Encoder / Similaridade) ---
def encontrar_principio_similar_bi(principio_procurado, lista_principios, threshold=0.98):
    principio_procurado_norm = normalizar(principio_procurado)
    
    for principio_lista in lista_principios:
        principio_lista_norm = normalizar(principio_lista)
        
        # 1. Match Exato
        if principio_procurado_norm == principio_lista_norm: 
            return True, 1.0
        
        # 2. Contenção (com trava de tamanho)
        if principio_procurado_norm in principio_lista_norm or principio_lista_norm in principio_procurado_norm:
            if abs(len(principio_procurado_norm) - len(principio_lista_norm)) <= 3: 
                return True, 0.99
        
        # 3. Similaridade Vetorial (Bi-Encoder)
        if modelo_bi:
            try:
                v1 = modelo_bi.encode(principio_procurado_norm, convert_to_tensor=True)
                v2 = modelo_bi.encode(principio_lista_norm, convert_to_tensor=True)
                sim = util.cos_sim(v1, v2)[0][0].item()
                if sim >= threshold: 
                    return True, sim
            except: 
                pass
    return False, 0


# --- LÓGICA 2: NÃO MEDICAMENTOS (Cross-Encoder / Interpretação) ---
def interpretar_descricao_cross_encoder(texto_produto, texto_regra, threshold=1.0):
    """
    Score do MMarco:
    < 0: Nada a ver
    > 0: Alguma relação
    > 1: Relação forte (Recomendado)
    """
    # Remove "exceto..." da regra para não confundir a interpretação positiva
    texto_regra_limpo = re.split(r'exceto|exceção|salvo', texto_regra, flags=re.IGNORECASE)[0].strip()
    
    par = [texto_regra_limpo, texto_produto]
    
    try:
        # CrossEncoder espera uma lista de pares
        score = modelo_cross.predict([par])[0]
        
        if score > threshold:
            return True, float(score)
    except Exception as e:
        print(f"Erro Cross-Encoder: {e}")
        
    return False, float(score)


def verificar_match_hibrido(item_dados, tabela, threshold_medicamento=0.98, threshold_descricao=1.0):
    matches_encontrados = []
    
    ncm_item_limpo = limpar_ncm(item_dados['ncm'])
    eh_medicamento_item = item_dados['eh_medicamento']
    
    for linha_tabela in tabela:
        if len(linha_tabela) < 6: continue
        
        # Colunas do Sheets
        item_tab = linha_tabela[0]
        anexo_tab = linha_tabela[1]
        desc_princ_tab = linha_tabela[2] # Texto da regra completo
        ncm_tab = linha_tabela[3]
        aliquota_tab = linha_tabela[4]
        tipo_tab = str(linha_tabela[5]).strip().lower()

        # --- 1. FILTRO DE TIPO ---
        match_tipo = False
        if eh_medicamento_item and "medicamento" in tipo_tab and "não" not in tipo_tab: 
            match_tipo = True
        elif not eh_medicamento_item and ("não medicamento" in tipo_tab or "nao medicamento" in tipo_tab): 
            match_tipo = True
            
        if not match_tipo: continue

        # --- 2. VALIDAÇÃO DE NCM (Unidirecional + Exceção) ---
        ncm_tab_limpo = limpar_ncm(ncm_tab)
        match_ncm = False
        
        if ncm_item_limpo and ncm_tab_limpo:
            # O Item (específico) deve começar com o NCM da Tabela (Genérico)
            if ncm_item_limpo.startswith(ncm_tab_limpo):
                match_ncm = True
                
                # Verifica exceções na descrição da tabela
                ncms_excluidos = extrair_ncms_excecao(desc_princ_tab)
                if ncms_excluidos:
                    for excluido in ncms_excluidos:
                        # Se o item cai na exceção
                        if ncm_item_limpo.startswith(excluido):
                            match_ncm = False
                            break
        
        if not match_ncm: continue

        # --- 3. VALIDAÇÃO DE CONTEÚDO (DIVIDIDA POR TIPO) ---
        match_conteudo = False
        detalhes_match = []

        if eh_medicamento_item:
            # === LÓGICA ANTIGA (Bi-Encoder) para Medicamentos ===
            principios_item = item_dados['principios']
            
            # Limpa regra para extrair princípios da tabela
            texto_limpo = re.split(r'exceto|exceção|salvo', str(desc_princ_tab), flags=re.IGNORECASE)[0]
            
            principios_tabela = []
            if "+" in texto_limpo:
                principios_tabela = [p.strip() for p in texto_limpo.split("+") if p.strip()]
            else:
                principios_tabela = [texto_limpo.strip()]
            
            # Regras estritas de quantidade e vazio
            if not principios_tabela or principios_tabela == ['']: continue
            if len(principios_item) != len(principios_tabela): continue

            # Match de Ida (Item -> Tabela)
            todos_item_na_tabela = True
            for pi in principios_item:
                enc, _ = encontrar_principio_similar_bi(pi, principios_tabela, threshold_medicamento)
                if not enc: todos_item_na_tabela = False; break
            
            # Match de Volta (Tabela -> Item)
            todos_tabela_no_item = True
            if todos_item_na_tabela:
                for pt in principios_tabela:
                    enc, _ = encontrar_principio_similar_bi(pt, principios_item, threshold_medicamento)
                    if not enc: todos_tabela_no_item = False; break
                    else: detalhes_match.append(pt)
            
            if todos_item_na_tabela and todos_tabela_no_item: match_conteudo = True

        else:
            # === NOVA LÓGICA (Cross-Encoder) para Não Medicamentos ===
            descricao_item = item_dados['descricao_completa']
            
            # Usa a IA pesada para interpretar se bate
            match_conteudo, score = interpretar_descricao_cross_encoder(descricao_item, desc_princ_tab, threshold_descricao)
            
            if match_conteudo:
                detalhes_match.append(f"{desc_princ_tab} (Score Cross: {score:.2f})")

        # --- SE PASSOU POR TUDO, GUARDA O MATCH ---
        if match_conteudo:
            matches_encontrados.append({
                'item_tabela': item_tab,
                'anexo_tabela': anexo_tab,
                'principio_tabela': desc_princ_tab,
                'aliquota_tabela': aliquota_tab,
                'detalhes': detalhes_match
            })

    return len(matches_encontrados) > 0, matches_encontrados


def lista_itens_excel():
    carregar_modelos()
    
    print("Buscando dados do Sheets...")
    tabela_completa = GeraTabelasSheetsIsencao.gera_tabela_sheets_isencao()
    
    caminho_arquivo = r"C:\rpa\Python\Analisa Itens Reforma Tributaria\Nova Analise\ANALISE.xlsx"
    workbook = load_workbook(caminho_arquivo)
    aba_ativa = workbook.active
    linha = 2
    
    print("Iniciando varredura...")
    
    for row in aba_ativa.iter_rows(min_row=2, values_only=True):
        # Mapeamento de colunas (Ajuste se necessário)
        nome_mercadoria = row[1] if row[1] else ""
        descricao_mercadoria = row[2] if row[2] else ""
        principios_ativos_str = row[3] if row[3] else ""
        ncm = row[5]
        analise_ia_excel = row[10]
        convenio_isencao = row[14]
        
        if analise_ia_excel is None or analise_ia_excel == "":
            retorno_ia = "NÃO"
            texto_match_str = ""
            indices_anexo_str = ""
            
            # --- DECISÃO: É Medicamento ou Não? ---
            eh_medicamento = True
            # Se não tem princípios ou NCMs tipicos de material, vira Não Medicamento
            if " - - " in str(principios_ativos_str) or not principios_ativos_str:
                eh_medicamento = False
            
            principios_lista = []
            descricao_completa = ""
            
            if eh_medicamento:
                # Prepara dados para Bi-Encoder
                principios_lista = [extrair_nome_principio(p).strip() for p in str(principios_ativos_str).split("|")]
            else:
                # Prepara dados para Cross-Encoder (Precisa de contexto rico)
                descricao_completa = f"Produto: {nome_mercadoria}. Descrição Técnica: {descricao_mercadoria}. NCM: {ncm}".strip()

            item_dados = {
                'ncm': ncm,
                'principios': principios_lista,
                'descricao_completa': descricao_completa,
                'eh_medicamento': eh_medicamento
            }
            
            # Chama o Match
            match_found, detalhes = verificar_match_hibrido(item_dados, tabela_completa)
            
            if match_found:
                retorno_ia = "SIM"
                lista_matchs = []
                lista_anexos = []
                for d in detalhes:
                    lista_matchs.append(str(d['principio_tabela']))
                    desc_anexo = f"{d['anexo_tabela']} - Item {d['item_tabela']} ({d['aliquota_tabela']})"
                    lista_anexos.append(desc_anexo)
                
                texto_match_str = " | ".join(lista_matchs)
                indices_anexo_str = " | ".join(lista_anexos)
            
            # Verifica Isenção
            if convenio_isencao and "sem isenção" in str(convenio_isencao).lower():
                isento_convenio = "NÃO"
            else:
                isento_convenio = "SIM"
            
            preenche_informacao_excel(aba_ativa, linha, retorno_ia, texto_match_str, indices_anexo_str, isento_convenio)
            
            if linha % 50 == 0:
                print(f"Processando linha {linha}...")
                workbook.save(caminho_arquivo)
                
        linha += 1
    
    workbook.save(caminho_arquivo)
    workbook.close()
    print("Processamento concluído!")


# Executar a função
if __name__ == "__main__":
    lista_itens_excel()