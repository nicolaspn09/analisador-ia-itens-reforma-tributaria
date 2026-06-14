import os
import shutil
from openpyxl import load_workbook



class GeraTabelaExcel:
    def __init__(self):
        pass


    def movimenta_arquivos():
        try:
            os.remove(r"C:\rpa\Python\Analisa Itens Reforma Tributaria\Nova Analise\ANALISE.xlsx")
        except:
            pass


        caminho_origem = r"C:\rpa\Python\Analisa Itens Reforma Tributaria\Nova Analise\Clear\ANALISE.xlsx"
        caminho_destino_completo = r"C:\rpa\Python\Analisa Itens Reforma Tributaria\Nova Analise\ANALISE.xlsx"
        shutil.copy(caminho_origem, caminho_destino_completo)


    def gera_tabela_excel(tabela):
        caminho_arquivo = r"C:\rpa\Python\Analisa Itens Reforma Tributaria\Nova Analise\ANALISE.xlsx"
        workbook = load_workbook(caminho_arquivo)
        sheet = workbook["Analise"]
        linha_inicial = 2

        for codigo_mercadoria, nome_mercadoria, descricao_mercadoria, principio_ativo, unidade_medida, ncm, ean_venda, ds_ministerio_saude, situacao_mercadoria, tipo_mercadoria,  mensagem_isencao in tabela:
            sheet.cell(row=linha_inicial, column=1).value = str(codigo_mercadoria)
            sheet.cell(row=linha_inicial, column=2).value = str(nome_mercadoria)
            sheet.cell(row=linha_inicial, column=3).value = str(descricao_mercadoria)
            sheet.cell(row=linha_inicial, column=4).value = str(principio_ativo)
            sheet.cell(row=linha_inicial, column=5).value = str(unidade_medida)
            sheet.cell(row=linha_inicial, column=6).value = str(ncm)
            sheet.cell(row=linha_inicial, column=7).value = str(ean_venda)
            sheet.cell(row=linha_inicial, column=8).value = str(ds_ministerio_saude)
            sheet.cell(row=linha_inicial, column=9).value = str(situacao_mercadoria)
            sheet.cell(row=linha_inicial, column=10).value = str(tipo_mercadoria)
            sheet.cell(row=linha_inicial, column=15).value = str(mensagem_isencao)

            linha_inicial += 1
            
        workbook.save(caminho_arquivo)
        workbook.close()