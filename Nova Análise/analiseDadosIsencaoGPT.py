import os
import re
import time
import unicodedata
import unidecode
from typing import List, Optional
from dotenv import load_dotenv, find_dotenv
from pydantic import BaseModel, Field
from langchain_openai import ChatOpenAI
from openpyxl import load_workbook
from sentence_transformers import SentenceTransformer, util
from geraTabelasSheetsIsencao import GeraTabelasSheetsIsencao

# --- CONFIGURAÇÃO ---
script_dir = os.path.dirname(os.path.abspath(__file__))
load_dotenv(find_dotenv(os.path.join(script_dir, '.env')))

os.environ["OPENAI_API_KEY"] = os.getenv("OPENAI_API_KEY")
MODEL_NAME = "gpt-4.1-nano" # Ajuste conforme disponibilidade

print("Carregando modelo local para Medicamentos...")
modelo_local = SentenceTransformer('paraphrase-multilingual-MiniLM-L12-v2')

# --- FUNÇÕES AUXILIARES ---
def normalizar(texto):
    if not texto: 
        return ""
    return unidecode.unidecode(str(texto).upper()).strip()

def limpar_ncm(ncm):
    """Remove pontos e retorna apenas números."""
    if not ncm: 
        return ""
    return "".join(filter(str.isdigit, str(ncm)))

def extrair_lista_ncms_da_regra(campo_ncm_tabela):
    """
    Transforma '9022.13 / 9022.14' em ['902213', '902214'].
    Lida com barras, 'ou' e espaços.
    """
    if not campo_ncm_tabela: 
        return []
    # Padroniza separadores
    texto = str(campo_ncm_tabela).lower().replace("/", " ").replace(" ou ", " ").replace(",", " ")
    # Quebra por espaço e limpa cada parte
    partes = texto.split()
    lista_final = []
    for p in partes:
        limpo = limpar_ncm(p)
        if limpo: # Só adiciona se não for vazio
            lista_final.append(limpo)
    return lista_final

def extrair_nome_principio(principio_completo):
    partes = principio_completo.split(" - ")
    if len(partes) >= 3:
        return partes[1].strip()
    return principio_completo.strip()

# --- LÓGICA LOCAL (MEDICAMENTOS) ---
def encontrar_principio_similar_local(principio_procurado, lista_principios, threshold=0.98):
    principio_procurado_norm = normalizar(principio_procurado)
    for principio_lista in lista_principios:
        principio_lista_norm = normalizar(principio_lista)
        if principio_procurado_norm == principio_lista_norm: 
            return True
        if (principio_procurado_norm in principio_lista_norm or principio_lista_norm in principio_procurado_norm):
            if abs(len(principio_procurado_norm) - len(principio_lista_norm)) <= 3: 
                return True
        try:
            v1 = modelo_local.encode(principio_procurado_norm, convert_to_tensor=True)
            v2 = modelo_local.encode(principio_lista_norm, convert_to_tensor=True)
            valor_analisar = util.cos_sim(v1, v2)[0][0].item()
            if valor_analisar >= threshold: 
                return True
        except:
            pass
    return False

def analisar_medicamento_local(item_dados, tabela_completa):
    ncm_item = limpar_ncm(item_dados['ncm'])
    principios_item = [extrair_nome_principio(p).strip() for p in str(item_dados['principios']).split("|")]
    
    for linha in tabela_completa:
        if len(linha) < 6: 
            continue
        if "não medicamento" in str(linha[5]).lower(): 
            continue 

        # --- NOVA LÓGICA DE NCM MÚLTIPLO ---
        ncms_regra = extrair_lista_ncms_da_regra(linha[3])
        match_ncm = False
        for ncm_r in ncms_regra:
            # Item começa com algum dos NCMs da regra?
            if ncm_item.startswith(ncm_r):
                match_ncm = True
                break
        
        if not match_ncm: 
            continue
        # -------------------------------------

        texto_limpo = re.split(r'exceto|exceção|salvo', str(linha[2]), flags=re.IGNORECASE)[0]
        principios_tab = [p.strip() for p in texto_limpo.split("+") if p.strip()]
        
        if not principios_tab or len(principios_item) != len(principios_tab): 
            continue

        match_ida = all(encontrar_principio_similar_local(p, principios_tab) for p in principios_item)
        match_volta = all(encontrar_principio_similar_local(p, principios_item) for p in principios_tab)

        if match_ida and match_volta:
            return {
                "match": True,
                "texto_coluna_L": linha[2],
                "anexo": f"{linha[1]} - Item {linha[0]}",
                "aliquota": str(linha[4])
            }
            
    return {"match": False}

# --- LÓGICA GPT (NÃO MEDICAMENTOS) ---
def formatar_regras_gpt(regras):
    texto = ""
    for i, r in enumerate(regras):
        texto += f"ID_{i}: [NCMs: {r[3]}] - Descrição: {r[2]}\n"
    return texto

# --- NOVA ESTRUTURA DE SAÍDA (Mais robusta) ---
class AnaliseTributaria(BaseModel):
    raciocinio: str = Field(description="Pense passo a passo. Compare a descrição técnica do item com a descrição da regra. Aponte discrepâncias.")
    match: bool = Field(description="Defina como True APENAS se o item for EXATAMENTE o descrito na regra. Na dúvida, False.")
    id_regra: Optional[str] = Field(default=None, description="O ID da regra que deu match exato.")

# --- LÓGICA GPT CORRIGIDA ---
def analisar_nao_medicamento_gpt(llm, item_dados, regras_candidatas):
    if not regras_candidatas:
        return {"match": False}

    texto_regras = formatar_regras_gpt(regras_candidatas)

    prompt_sys = """
    Você é um Auditor Fiscal Técnico. Sua análise deve ser baseada na FUNÇÃO EXATA do produto.
    
    PASSO A PASSO DA ANÁLISE:
    1. **Identifique o Produto Real:** Decodifique abreviações e marcas (ex: "CLINENSOL" -> Solução de Limpeza de Lentes).
    2. **Compare a Finalidade de Uso (CRUCIAL):** - O produto serve PARA A MESMA COISA que a regra descreve?
       - "Líquido para limpar lentes" (Higiene externa) NÃO É "Solução para Diálise" (Procedimento intra-corpóreo/sanguíneo).
       - Se a função for diferente, REJEITE, mesmo que a consistência física (líquido, creme) seja igual.
    3. **Natureza Química:** Produtos de limpeza/higiene (mesmo vendidos em farmácia) não são dispositivos médicos complexos (como enxertos ou soluções de diálise).

    EXEMPLOS PARA CALIBRAGEM:
    - Item: "CLINENSOL LIQ", Regra: "Concentrados para Diálise".
      -> Raciocínio: Clinensol serve para limpar lentes de contato. Diálise é filtragem sanguínea. Funções distintas. -> MATCH: FALSE.
    
    - Item: "CR SORRISO", Regra: "Dentifrícios". 
      -> Raciocínio: Sorriso é marca de pasta de dente. Pasta de dente = Dentifrício. -> MATCH: TRUE.

    - Item: "ACETONA", Regra: "Preservativos".
      -> Raciocínio: Solvente vs Contraceptivo. -> MATCH: FALSE.

    VEREDITO:
    Retorne MATCH=TRUE apenas se houver identidade total de função e descrição. Na dúvida, ou se não conhecer a marca, MATCH=FALSE.
    """
    
    prompt_user = f"""
    ITEM PARA ANÁLISE:
    Descrição Comercial: {item_dados['nome']}
    Descrição Adicional: {item_dados['descricao']}
    NCM Informado: {item_dados['ncm']}
    
    REGRAS CANDIDATAS:
    {texto_regras}
    
    Analise a FINALIDADE DO USO no raciocínio.
    """
    
    structured_llm = llm.with_structured_output(AnaliseTributaria)
    
    try:
        resultado = structured_llm.invoke([("system", prompt_sys), ("user", prompt_user)])
        
        if resultado.match and resultado.id_regra:
            try:
                index = int(resultado.id_regra.replace("ID_", ""))
                if 0 <= index < len(regras_candidatas):
                    regra = regras_candidatas[index]
                    
                    if not item_dados['nome'] and not item_dados['descricao']: return {"match": False}

                    return {
                        "match": True,
                        "texto_coluna_L": f"Isenção: {regra[2]}",
                        "anexo": f"{regra[1]} - Item {regra[0]}",
                        "aliquota": str(regra[4]),
                        "raciocinio_gpt": resultado.raciocinio
                    }
            except ValueError:
                pass
                
        return {"match": False}
    except Exception as e:
        print(f"Erro GPT: {e}")
        return {"match": False}

# --- FUNÇÃO PRINCIPAL ---
def lista_itens_excel():
    print(f"Iniciando GPT ({MODEL_NAME})...")
    llm = ChatOpenAI(model=MODEL_NAME, temperature=0) # Temperature padrão

    print("Carregando Tabela de Regras...")
    tabela_completa = GeraTabelasSheetsIsencao.gera_tabela_sheets_isencao()
    
    caminho = r"C:\rpa\Python\Analisa Itens Reforma Tributaria\Nova Analise\ANALISE.xlsx"
    wb = load_workbook(caminho)
    aba = wb.active
    linha = 2
    
    print("Iniciando varredura...")
    
    for row in aba.iter_rows(min_row=2, values_only=True):
        nome, desc = (row[1] or ""), (row[2] or "")
        princ_str, ncm = (row[3] or ""), row[5]
        tipo_produto = (row[9] or "")
        analise_existente = row[10]
        convenio = row[14]
        
        if not analise_existente:
            item_dados = {"nome": nome, "descricao": desc, "ncm": ncm, "principios": princ_str}
            
            eh_medicamento = True
            sem_principio = " - - " in str(princ_str) or not princ_str
            tipo_nao_medicamento = "não" in str(tipo_produto).lower() or "nao" in str(tipo_produto).lower()
            
            if sem_principio or tipo_nao_medicamento:
                eh_medicamento = False
            
            resultado_final = {"match": False}
            
            if eh_medicamento:
                resultado_final = analisar_medicamento_local(item_dados, tabela_completa)
            else:
                # --- FILTRO PRÉVIO MELHORADO ---
                ncm_item_limpo = limpar_ncm(ncm)
                candidatos = []
                
                # Só tenta buscar regra se o NCM do item tiver pelo menos 4 dígitos.
                # NCMs de 2 dígitos geram falsos positivos demais.
                if ncm_item_limpo and len(ncm_item_limpo) >= 4:
                    for i, reg in enumerate(tabela_completa): # Use enumerate para debugging se precisar
                        if len(reg) < 6: 
                            continue
                        
                        # Filtra linhas que são explicitamente só medicamentos
                        # Ajuste conforme sua tabela: se a coluna 5 diz "Medicamento", ignorar.
                        if "não medicamento" not in str(reg[5]).lower(): 
                             # Se sua tabela tem itens mistos, revise essa lógica. 
                             # Mas itens médicos geralmente não devem bater com químicos puros.
                             pass 

                        ncms_regra_lista = extrair_lista_ncms_da_regra(reg[3])
                        
                        match_ncm_candidato = False
                        for ncm_r in ncms_regra_lista:
                            # INVERSÃO DE LOGICA:
                            # O NCM da Regra (ncm_r) é que manda.
                            # Se a regra é "3004", e o item é "30049099", OK.
                            # Se a regra é "30049099" e o item é "3004", NÃO OK (Item genérico demais para regra especifica).
                            
                            if len(ncm_r) < 4: continue # Ignora regras com NCMs muito curtos (capitulos inteiros)

                            if ncm_item_limpo.startswith(ncm_r):
                                match_ncm_candidato = True
                                break
                        
                        if match_ncm_candidato:
                            candidatos.append(reg)
                
                # Só chama o GPT se tiver candidatos REAIS
                if candidatos:
                    resultado_final = analisar_nao_medicamento_gpt(llm, item_dados, candidatos)
                else:
                    resultado_final = {"match": False}

                if resultado_final["match"]:
                    print(f"GPT Match: {nome} -> {resultado_final['anexo']}")

            # --- PREENCHIMENTO ---
            aba.cell(row=linha, column=11).value = "SIM" if resultado_final["match"] else "NÃO"
            
            val_L = resultado_final.get("texto_coluna_L", "") if resultado_final["match"] else ""
            aba.cell(row=linha, column=12).value = val_L
            
            val_M = f"{resultado_final.get('anexo', '')} ({resultado_final.get('aliquota', '')})" if resultado_final["match"] else ""
            aba.cell(row=linha, column=13).value = val_M
            
            isento_cv = "NÃO" if convenio and "sem isenção" in str(convenio).lower() else "SIM"
            aba.cell(row=linha, column=14).value = isento_cv
            
            if linha % 50 == 0:
                print(f"Salvando linha {linha}...")
                wb.save(caminho)
                
        linha += 1
    
    wb.save(caminho)
    wb.close()
    print("Concluído!")

if __name__ == "__main__":
    lista_itens_excel()